from __future__ import annotations

import io
from fastapi import FastAPI, File, UploadFile, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ifc_parser import parse_ifc
from rab_calculator import calculate_rab, total_rab, DEFAULT_HARGA

app = FastAPI(title="IFCost API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://ifcost-web.vercel.app",
        "http://localhost:5173",
        "http://localhost:4173",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def root():
    return {"status": "ok", "version": "1.0.0"}


@app.post("/api/analyze")
async def analyze(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".ifc"):
        raise HTTPException(status_code=400, detail="File harus berformat .ifc")

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Ukuran file maksimum 50MB")

    try:
        result = parse_ifc(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Gagal memproses file IFC: {str(e)}")

    return {
        "filename": file.filename,
        **result,
    }


@app.post("/api/export/excel")
async def export_excel(payload: dict = Body(...)):
    """Export data QTO + RAB ke file Excel."""
    try:
        qto = payload.get("qto", {})
        spaces = payload.get("spaces", [])
        totals = payload.get("totals", {})
        harga_satuan = payload.get("harga_satuan", DEFAULT_HARGA)
        filename = payload.get("filename", "model")

        rab_items = calculate_rab(totals, harga_satuan)
        total = total_rab(rab_items)

        wb = openpyxl.Workbook()

        # --- Sheet 1: Ringkasan Elemen ---
        ws1 = wb.active
        ws1.title = "Ringkasan_Elemen"
        _write_summary_sheet(ws1, totals, payload.get("summary", {}))

        # --- Sheet 2: QTO Dinding ---
        ws2 = wb.create_sheet("QTO_Dinding")
        _write_wall_sheet(ws2, qto.get("walls", []))

        # --- Sheet 3: QTO Lantai ---
        ws3 = wb.create_sheet("QTO_Lantai")
        _write_slab_sheet(ws3, qto.get("slabs", []))

        # --- Sheet 4: QTO Kolom ---
        ws4 = wb.create_sheet("QTO_Kolom")
        _write_column_sheet(ws4, qto.get("columns", []), qto.get("beams", []))

        # --- Sheet 5: RAB ---
        ws5 = wb.create_sheet("Estimasi_Biaya_RAB")
        _write_rab_sheet(ws5, rab_items, total)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = filename.replace(".ifc", "").replace(" ", "_")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="ifcost_{safe_name}.xlsx"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal membuat Excel: {str(e)}")


# ── Excel helpers ──────────────────────────────────────────────────────────────

BLUE = "FF3B82F6"
GREEN = "FF22C55E"
HEADER_BG = "FF1E40AF"
HEADER_FG = "FFFFFFFF"
ALT_BG = "FFF1F3F5"

header_font = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
header_fill = PatternFill("solid", fgColor=HEADER_BG)
header_align = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color="FFE5E7EB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_row(ws, cols: list[str], row: int = 1):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[row].height = 22


def _data_row(ws, values: list, row: int, alt: bool = False):
    fill = PatternFill("solid", fgColor=ALT_BG) if alt else None
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = border
        if fill:
            cell.fill = fill


def _write_summary_sheet(ws, totals: dict, summary: dict):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    _header_row(ws, ["Metrik", "Nilai"])
    data = [
        ("Total Dinding (elemen)", summary.get("IfcWall", 0) + summary.get("IfcWallStandardCase", 0)),
        ("Total Lantai (elemen)", summary.get("IfcSlab", 0)),
        ("Total Kolom (elemen)", summary.get("IfcColumn", 0)),
        ("Total Balok (elemen)", summary.get("IfcBeam", 0)),
        ("Total Jendela (elemen)", summary.get("IfcWindow", 0)),
        ("Total Pintu (elemen)", summary.get("IfcDoor", 0)),
        ("Volume Dinding (m³)", totals.get("wall_volume", 0)),
        ("Luas Dinding (m²)", totals.get("wall_area", 0)),
        ("Luas Lantai Slab (m²)", totals.get("slab_area", 0)),
        ("Volume Kolom (m³)", totals.get("column_volume", 0)),
        ("Volume Balok (m³)", totals.get("beam_volume", 0)),
        ("Total Luas Lantai (m²)", totals.get("floor_area", 0)),
    ]
    for i, (label, val) in enumerate(data, 2):
        _data_row(ws, [label, val], i, alt=i % 2 == 0)


def _write_wall_sheet(ws, walls: list):
    cols = ["GlobalId", "Nama", "Tipe", "Panjang (m)", "Tinggi (m)", "Luas Bersih (m²)", "Volume Bersih (m³)", "Volume Bruto (m³)"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    for i, w in enumerate(walls, 2):
        _data_row(ws, [
            w.get("GlobalId"), w.get("Name"), w.get("Type"),
            w.get("Length"), w.get("Height"),
            w.get("NetSideArea"), w.get("NetVolume"), w.get("GrossVolume"),
        ], i, alt=i % 2 == 0)


def _write_slab_sheet(ws, slabs: list):
    cols = ["GlobalId", "Nama", "PredefinedType", "Luas Bruto (m²)", "Luas Bersih (m²)", "Volume (m³)"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    for i, s in enumerate(slabs, 2):
        _data_row(ws, [
            s.get("GlobalId"), s.get("Name"), s.get("PredefinedType"),
            s.get("GrossArea"), s.get("NetArea"), s.get("Volume"),
        ], i, alt=i % 2 == 0)


def _write_column_sheet(ws, columns: list, beams: list):
    ws.title = "QTO_Kolom_Balok"
    cols = ["Tipe Elemen", "GlobalId", "Nama", "Panjang (m)", "Luas Penampang (m²)", "Volume (m³)"]
    _header_row(ws, cols)
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    row = 2
    for el in columns:
        _data_row(ws, ["Kolom", el.get("GlobalId"), el.get("Name"), el.get("Length"), el.get("CrossSectionArea"), el.get("Volume")], row, alt=row % 2 == 0)
        row += 1
    for el in beams:
        _data_row(ws, ["Balok", el.get("GlobalId"), el.get("Name"), el.get("Length"), el.get("CrossSectionArea"), el.get("Volume")], row, alt=row % 2 == 0)
        row += 1


def _write_rab_sheet(ws, items: list, total: float):
    cols = ["Pekerjaan", "Volume", "Satuan", "Material", "Kebutuhan", "Harga Satuan (Rp)", "Subtotal (Rp)"]
    _header_row(ws, cols)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    row = 2
    for item in items:
        _data_row(ws, [
            item["pekerjaan"],
            item.get("volume"),
            item.get("satuan_volume"),
            item["material"],
            item["kebutuhan"],
            item["harga_satuan"],
            item["subtotal"],
        ], row)
        row += 1
        for sub in item.get("sub_items", []):
            _data_row(ws, [
                f"  ↳ {sub['material']}", None, sub["satuan"],
                sub["material"], sub["kebutuhan"],
                sub["harga_satuan"], sub["subtotal"],
            ], row, alt=True)
            row += 1

    # Total row
    total_cell = ws.cell(row=row, column=7, value=total)
    total_cell.font = Font(bold=True, color=GREEN, name="Calibri", size=12)
    ws.cell(row=row, column=1, value="TOTAL ESTIMASI RAB").font = Font(bold=True, name="Calibri", size=12)
